# Filename: Excel_graph_create.py
# Purpose:  For excel graph creating
# Author: Jack Chang

import openpyxl
import pandas as pd
from openpyxl.chart import LineChart, Reference

x = int(input("1. Storage\n2. Compute\n3. Storage_all\nPlease enter 1 or 2 or 3 to proceed: "))

def Storage():
    w = str(input("Please enter csv file name: "))
    z = str(input("Please enter xlsx file name: "))
    # ================Transform CSV to xlsx================
    workbook = pd.read_csv(("%s.csv") % w)
    workbook.to_excel(("%s.xlsx") % z, index=False,header=True)

    # ================Load xlsx file================
    wb = openpyxl.load_workbook(('%s.xlsx') % z)
    ws = wb.active
    
    # ================Create sheet================
    ws.title = (('%s') % z)
    ws2 = wb.create_sheet(index=1, title='Graph')

    # ================Remove disqualified inlet & SSD value================
    for i in range(ws.max_row, 1, -1):  # row 索引值
        delete_row = False
        for j in range(12, 13):  # column 索引值
            w = ws.cell(column=j, row=i).value  # 宣告w為儲存格內容值
            if (w is None) or (w < 43) or (w >= 48):  # 如果w值為空值, 或小於43, 或大於等於48者則刪除該列
                delete_row = True
        if delete_row:
            ws.delete_rows(i)
        delete_row = False
        for j in range(13, 29):  # column 索引值
            w = ws.cell(column=j, row=i).value  # 宣告w為儲存格內容值
            if (w is None) or (w < 45) or (w > 60):  # 如果w值為空值, 或小於45, 或大於60者則刪除該列
                delete_row = True
        if delete_row:
            ws.delete_rows(i)

    # ================Set title name, height, width, legend position, column and row number of graph================
    for num in range(1,6):
        globals()['chart%s' % num] = LineChart()
        if num == 1:
            chart1.title = 'FAN_RATE'
            chart1.height = 10
            chart1.width = 18
            chart1.legend.position = 'b'
            data = Reference(ws, min_col= 3, max_col= 9, min_row= 1, max_row= ws.max_row)
        if num == 2:
            chart2.title = 'PCIE_SWITCH_TEMP'
            chart2.height = 10
            chart2.width = 18
            chart2.legend.position = 'b'
            chart2.y_axis.scaling.max = 90
            data = Reference(ws, min_col= 10, max_col= 11, min_row= 1, max_row= ws.max_row)
        if num == 3:
            chart3.title = 'HSC_IN_POWER'
            chart3.height = 10
            chart3.width = 18
            chart3.legend.position = 'b'
            chart3.y_axis.scaling.max = 400
            data = Reference(ws, min_col= 2, max_col= 2, min_row= 1, max_row= ws.max_row)
        if num == 4:
            chart4.title = 'SYS_INLET_TEMP'
            chart4.height = 10
            chart4.width = 18
            chart4.legend.position = 'b'
            chart4.y_axis.scaling.min = 40
            chart4.y_axis.scaling.max = 50
            data = Reference(ws, min_col= 12, max_col= 12, min_row= 1, max_row= ws.max_row)
        if num == 5:
            chart5.title = 'SSD_TEMP'
            chart5.height = 10
            chart5.width = 18
            chart5.legend.position = 'b'
            chart5.y_axis.scaling.max = 80
            data = Reference(ws, min_col= 13, max_col= 28, min_row= 1, max_row= ws.max_row)
        globals()['chart%s' % num].add_data(data, titles_from_data=True)
        # ================Set x-axis of graph data range================
        xtitle = Reference(ws, min_col= 1, min_row= 2, max_row= ws.max_row)
        globals()['chart%s' % num].set_categories(xtitle)
        # ================Set location of graph & in which worksheet================
        if num == 1:
            ws2.add_chart(chart1, 'B2')
        if num == 2:
            ws2.add_chart(chart2, 'M2')
        if num == 3:
            ws2.add_chart(chart3, 'X2')
        if num == 4:
            ws2.add_chart(chart4, 'B22')
        if num == 5:
            ws2.add_chart(chart5, 'M22')
    wb.save(('%s.xlsx') % z)

def Compute():
    # ================Transform CSV to xlsx================
    workbook = pd.read_csv("Compute_all.csv")
    workbook.to_excel("Compute_all.xlsx", index=False,header=True)
    
    # ================Load xlsx file================
    wb = openpyxl.load_workbook('Compute_all.xlsx')
    ws = wb.active
    
    # ================Create sheet================
    ws.title = 'Inlet_value'
    ws2 = wb.create_sheet(index=1, title='Graph')

    # ================Remove disqualified inlet value================
    for i in range(ws.max_row, 1, -1):  # row 索引值
        delete_row = False
        for j in range(2, 7):  # column 索引值
            w = ws.cell(column=j, row=i).value  # 宣告w為儲存格內容值
            if (w is None) or (w < 43) or (w >= 48):  # 如果w值為空值, 或小於43, 或大於等於48者則刪除該列
                delete_row = True
        if delete_row:
            ws.delete_rows(i)

    # ================Set title name, height, width, legend position, column and row number of graph================
    for num in range(1,3):
        globals()['chart%s' % num] = LineChart()
        if num == 1:
            chart1.title = 'COMPUTE_TEMP_INLET'
            chart1.height = 10
            chart1.width = 18
            chart1.legend.position = 'b'
            chart1.y_axis.scaling.min = 40
            chart1.y_axis.scaling.max = 50
            data = Reference(ws, min_col= 2, max_col= 7, min_row= 1, max_row= ws.max_row)
        if num == 2:
            chart2.title = 'COMPUTE_HSC0_Input_Power'
            chart2.height = 10
            chart2.width = 18
            chart2.legend.position = 'b'
            chart2.y_axis.scaling.max = 500
            data = Reference(ws, min_col= 8, max_col= 13, min_row= 1, max_row= ws.max_row)
        globals()['chart%s' % num].add_data(data, titles_from_data=True)
        # ================Set x-axis of graph data range================
        xtitle = Reference(ws, min_col= 1, min_row= 2, max_row= ws.max_row)
        globals()['chart%s' % num].set_categories(xtitle)
        # ================Set location of graph & in which worksheet================
        if num == 1:
            ws2.add_chart(chart1, 'B2')
        if num == 2:
            ws2.add_chart(chart2, 'M2')
    wb.save('Compute_all.xlsx')

def Storage_all():
    w = str(input("Please enter csv file name: "))
    # ================Transform CSV to xlsx================
    workbook = pd.read_csv(("%s.csv") % w)
    workbook.to_excel("Storage_all.xlsx", index=False, header=True)

    # ================Load xlsx file================
    wb = openpyxl.load_workbook("Storage_all.xlsx")
    ws = wb.active

    # ================Create sheet================
    ws.title = 'Storage_all'
    ws2 = wb.create_sheet(index=1, title='Graph')

    # ================Remove disqualified inlet & SSD value================
    for i in range(ws.max_row, 1, -1):  # row 索引值
        delete_row = False
        for j in range(2, 5):  # column 索引值
            w = ws.cell(column=j, row=i).value  # 宣告w為儲存格內容值
            if (w is None) or (w < 43) or (w >= 48):  # 如果w值為空值, 或小於43, 或大於等於48者則刪除該列
                delete_row = True
        if delete_row:
            ws.delete_rows(i)
        delete_row = False
        for j in range(5, 53):  # column 索引值
            w = ws.cell(column=j, row=i).value  # 宣告w為儲存格內容值
            if (w is None) or (w < 45) or (w > 60):  # 如果w值為空值, 或小於45, 或大於60者則刪除該列
                delete_row = True
        if delete_row:
            ws.delete_rows(i)

    # ================Set title name, height, width, legend position, column and row number of graph================
    for num in range(1, 18):
        globals()['chart%s' % num] = LineChart()
        if num == 1:
            chart1.title = 'SYS_INLET_TEMP'
            chart1.height = 10
            chart1.width = 18
            chart1.legend.position = 'b'
            chart1.y_axis.scaling.min = 40
            chart1.y_axis.scaling.max = 50
            data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
        if num == 2:
            chart2.title = 'SSD1'
            chart2.height = 10
            chart2.width = 18
            chart2.legend.position = 'b'
            chart2.y_axis.scaling.max = 80
            data = Reference(ws, min_col=5, max_col=7, min_row=1, max_row=ws.max_row)
        if num == 3:
            chart3.title = 'SSD2'
            chart3.height = 10
            chart3.width = 18
            chart3.legend.position = 'b'
            chart3.y_axis.scaling.max = 80
            data = Reference(ws, min_col=8, max_col=10, min_row=1, max_row=ws.max_row)
        if num == 4:
            chart4.title = 'SSD3'
            chart4.height = 10
            chart4.width = 18
            chart4.legend.position = 'b'
            chart4.y_axis.scaling.max = 80
            data = Reference(ws, min_col=11, max_col=13, min_row=1, max_row=ws.max_row)
        if num == 5:
            chart5.title = 'SSD4'
            chart5.height = 10
            chart5.width = 18
            chart5.legend.position = 'b'
            chart5.y_axis.scaling.max = 80
            data = Reference(ws, min_col=14, max_col=16, min_row=1, max_row=ws.max_row)
        if num == 6:
            chart6.title = 'SSD5'
            chart6.height = 10
            chart6.width = 18
            chart6.legend.position = 'b'
            chart6.y_axis.scaling.max = 80
            data = Reference(ws, min_col=17, max_col=19, min_row=1, max_row=ws.max_row)
        if num == 7:
            chart7.title = 'SSD6'
            chart7.height = 10
            chart7.width = 18
            chart7.legend.position = 'b'
            chart7.y_axis.scaling.max = 80
            data = Reference(ws, min_col=20, max_col=22, min_row=1, max_row=ws.max_row)
        if num == 8:
            chart8.title = 'SSD7'
            chart8.height = 10
            chart8.width = 18
            chart8.legend.position = 'b'
            chart8.y_axis.scaling.max = 80
            data = Reference(ws, min_col=23, max_col=25, min_row=1, max_row=ws.max_row)
        if num == 9:
            chart9.title = 'SSD8'
            chart9.height = 10
            chart9.width = 18
            chart9.legend.position = 'b'
            chart9.y_axis.scaling.max = 80
            data = Reference(ws, min_col=26, max_col=28, min_row=1, max_row=ws.max_row)
        if num == 10:
            chart10.title = 'SSD9'
            chart10.height = 10
            chart10.width = 18
            chart10.legend.position = 'b'
            chart10.y_axis.scaling.max = 80
            data = Reference(ws, min_col=29, max_col=31, min_row=1, max_row=ws.max_row)
        if num == 11:
            chart11.title = 'SSD10'
            chart11.height = 10
            chart11.width = 18
            chart11.legend.position = 'b'
            chart11.y_axis.scaling.max = 80
            data = Reference(ws, min_col=32, max_col=34, min_row=1, max_row=ws.max_row)
        if num == 12:
            chart12.title = 'SSD11'
            chart12.height = 10
            chart12.width = 18
            chart12.legend.position = 'b'
            chart12.y_axis.scaling.max = 80
            data = Reference(ws, min_col=35, max_col=37, min_row=1, max_row=ws.max_row)
        if num == 13:
            chart13.title = 'SSD12'
            chart13.height = 10
            chart13.width = 18
            chart13.legend.position = 'b'
            chart13.y_axis.scaling.max = 80
            data = Reference(ws, min_col=38, max_col=40, min_row=1, max_row=ws.max_row)
        if num == 14:
            chart14.title = 'SSD13'
            chart14.height = 10
            chart14.width = 18
            chart14.legend.position = 'b'
            chart14.y_axis.scaling.max = 80
            data = Reference(ws, min_col=41, max_col=43, min_row=1, max_row=ws.max_row)
        if num == 15:
            chart15.title = 'SSD14'
            chart15.height = 10
            chart15.width = 18
            chart15.legend.position = 'b'
            chart15.y_axis.scaling.max = 80
            data = Reference(ws, min_col=44, max_col=46, min_row=1, max_row=ws.max_row)
        if num == 16:
            chart16.title = 'SSD15'
            chart16.height = 10
            chart16.width = 18
            chart16.legend.position = 'b'
            chart16.y_axis.scaling.max = 80
            data = Reference(ws, min_col=47, max_col=49, min_row=1, max_row=ws.max_row)
        if num == 17:
            chart17.title = 'SSD16'
            chart17.height = 10
            chart17.width = 18
            chart17.legend.position = 'b'
            chart17.y_axis.scaling.max = 80
            data = Reference(ws, min_col=50, max_col=52, min_row=1, max_row=ws.max_row)
        globals()['chart%s' % num].add_data(data, titles_from_data=True)
        # ================Set x-axis of graph data range================
        xtitle = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        globals()['chart%s' % num].set_categories(xtitle)
        # ================Set location of graph & in which worksheet================
        if num == 1:
            ws2.add_chart(chart1, 'B2')
        if num == 2:
            ws2.add_chart(chart2, 'B22')
        if num == 3:
            ws2.add_chart(chart3, 'M2')
        if num == 4:
            ws2.add_chart(chart4, 'M22')
        if num == 5:
            ws2.add_chart(chart5, 'X2')
        if num == 6:
            ws2.add_chart(chart6, 'x22')
        if num == 7:
            ws2.add_chart(chart7, 'AI2')
        if num == 8:
            ws2.add_chart(chart8, 'AI22')
        if num == 9:
            ws2.add_chart(chart9, 'AT2')
        if num == 10:
            ws2.add_chart(chart10, 'AT22')
        if num == 11:
            ws2.add_chart(chart11, 'BE2')
        if num == 12:
            ws2.add_chart(chart12, 'BE22')
        if num == 13:
            ws2.add_chart(chart13, 'BP2')
        if num == 14:
            ws2.add_chart(chart14, 'BP22')
        if num == 15:
            ws2.add_chart(chart15, 'CA2')
        if num == 16:
            ws2.add_chart(chart16, 'CA22')
        if num == 17:
            ws2.add_chart(chart17, 'CL2')
    wb.save("Storage_all.xlsx")

if x == 1:
    Storage()

if x == 2:
    Compute()

if x == 3:
    Storage_all()